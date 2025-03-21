import os
import logging
import warnings

from openpyxl import load_workbook

from .extractor import Extractor

logger = logging.getLogger(__name__)
logger.setLevel(os.environ.get("LOG_LEVEL", "INFO"))


class ExcelExtractor(Extractor):
    file_types = ("xlsx",)

    _valid_cell_types = ["s", "n", "b", "inlineStr", "str"]

    def _extract(self, file: bytes) -> str:
        file_path = self._write_file(file, extension="xlsx")
        output = ""

        # Open workbook read-only to avoid massive slowdown while iterating through cells
        logger.debug("Opening workbook read-only to extract cell text.")
        workbook = load_workbook(file_path, read_only=True)
        for sheet in workbook:
            logger.debug("Extracting cell text from worksheet \"%s\".", sheet.title)
            for row in sheet:
                for cell in row:
                    if cell.value and cell.data_type in self._valid_cell_types:
                        output += f" {str(cell.value)}"
        workbook.close()

        # Reopen workbook read-write so that the "_images" attribute is available
        logger.debug("Opening workbook read-write to extract images.")
        warnings.filterwarnings("ignore", category=UserWarning)  # Hide unnecessary warning about the data validation extension
        workbook = load_workbook(file_path)
        warnings.resetwarnings()
        for sheet in workbook:
            logger.debug("Extracting images from worksheet \"%s\".", sheet.title)
            for i, image in enumerate(sheet._images):
                file_name = f"{sheet.title}_image{i}.{image.format}"
                output += f" {self._extract_embedded(file_name, image._data())}"
        workbook.close()

        return output
